# %% [markdown]
# # Notebook for converting the grocery upc spreadsheet into a database migration
# %%
import openpyxl

SPREADSHEET_FILEPATH = r'.\Grocery_UPC_Database.xlsx'
spreadsheet = openpyxl.load_workbook(SPREADSHEET_FILEPATH)
# %%
# %%
ws = spreadsheet.active
x = next(ws.iter_rows())
x
# %%
x[3].value
# %%
pickles = []
for row in ws.iter_rows():
    if len(row) == 5:
        (maker, product) = list(map(lambda x: x.value, row[3:5]))
        screaming_product = product.upper()
        keywords = ['PICKLE', 'KIMCHI', 'SAUERKRAUT']
        for keyword in keywords:
            if keyword in screaming_product:
                print(f'Added {product} ({maker})')
                pickles.append((maker, product))
                break
    else:
        print(f'Row length: {len(row)}')
# %%
len(pickles)
# %%
[y for (x,y) in pickles if 'VLASIC' in y.upper()]
# %%
# now add values to data migration
# https://django.readthedocs.io/en/stable/topics/migrations.html#data-migrations
import pyperclip

pyperclip.copy('\n'.join(f'("{maker}", "{product}")' for (maker, product) in pickles))

# %%
